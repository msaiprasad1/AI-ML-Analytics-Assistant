from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.line_chart import LineChart
from openpyxl.drawing.image import Image
from datetime import datetime
import pandas as pd
import os

def write_results_to_excel(file_path, results):
    """
    Write analysis results to organized Excel sheets with formatting and charts
    """
    try:
        # Try to load existing workbook, create new one if it doesn't exist
        try:
            wb = load_workbook(file_path)
        except:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        print(f"Writing {len(results)} result sheets to Excel...")
        
        for sheet_name, df in results.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                # Convert data types to be Excel-compatible
                df_clean = df.copy()
                
                # Reset index to avoid index-related issues
                df_clean = df_clean.reset_index(drop=True)
                
                # Convert problematic data types
                for col in df_clean.columns:
                    if df_clean[col].dtype == 'int64':
                        df_clean[col] = df_clean[col].astype('int32')
                    elif df_clean[col].dtype == 'float64':
                        df_clean[col] = df_clean[col].astype('float32')
                    elif df_clean[col].dtype == 'object':
                        # Convert object columns to string to avoid issues
                        df_clean[col] = df_clean[col].astype(str)
                
                # Ensure all values are Excel-compatible
                df_clean = df_clean.fillna('')
                
                # Create sheet with timestamp
                ws_name = f"{sheet_name}_{timestamp}"
                ws = wb.create_sheet(ws_name)
                
                # Write data to sheet
                for r in dataframe_to_rows(df_clean, index=False, header=True):
                    ws.append(r)
                
                # Apply formatting
                format_excel_sheet(ws, df_clean, header_font, header_fill, border)
                
                # Add charts for specific sheets
                add_charts_to_sheet(ws, df_clean, sheet_name)
                
                print(f"Created sheet: {ws_name}")
        
        # Save the workbook
        wb.save(file_path)
        print(f"Excel file saved: {file_path}")
        
        return f"Successfully wrote {len(results)} sheets to Excel"
        
    except Exception as e:
        print(f"Error writing to Excel: {str(e)}")
        return f"Error: {str(e)}"

def format_excel_sheet(ws, df, header_font, header_fill, border):
    """
    Apply formatting to Excel sheet
    """
    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def add_charts_to_sheet(ws, df, sheet_name):
    """
    Add appropriate charts based on sheet content
    """
    try:
        # Correlation Matrix Chart
        if 'Correlation' in sheet_name and len(df.columns) > 1:
            add_correlation_chart(ws, df)
        
        # Model Performance Chart
        elif 'Model_Comparison' in sheet_name:
            add_model_comparison_chart(ws, df)
        
        # Feature Importance Chart
        elif 'Feature_Importance' in sheet_name or 'Random_Forest_Importance' in sheet_name:
            add_feature_importance_chart(ws, df)
        
        # Predictions vs Actual Chart
        elif 'Predictions_vs_Actual' in sheet_name:
            add_predictions_chart(ws, df)
        
        # Missing Values Chart
        elif 'Missing_Values' in sheet_name and not df.empty:
            add_missing_values_chart(ws, df)
            
    except Exception as e:
        print(f"⚠️ Could not add chart to {sheet_name}: {str(e)}")

def add_correlation_chart(ws, df):
    """
    Add correlation heatmap-style visualization
    """
    try:
        # Create a simple bar chart for top correlations
        if 'High_Correlations' in str(df.columns):
            chart = LineChart()
            chart.title = "Top Correlations"
            chart.style = 13
            chart.y_axis.title = 'Correlation Value'
            chart.x_axis.title = 'Feature Pairs'
            
            # Add data
            data = Reference(ws, min_col=3, min_row=2, max_row=min(ws.max_row, 11), max_col=3)
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 11))
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            ws.add_chart(chart, f"F{ws.max_row + 3}")
    except:
        pass

def add_model_comparison_chart(ws, df):
    """
    Add model comparison chart
    """
    try:
        chart = LineChart()
        chart.title = "Model Performance Comparison"
        chart.style = 13
        chart.y_axis.title = 'R² Score'
        chart.x_axis.title = 'Models'
        
        # Add R² scores
        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"F{ws.max_row + 3}")
    except:
        pass

def add_feature_importance_chart(ws, df):
    """
    Add feature importance chart
    """
    try:
        chart = LineChart()
        chart.title = "Feature Importance"
        chart.style = 13
        chart.y_axis.title = 'Importance'
        chart.x_axis.title = 'Features'
        
        # Add importance values
        importance_col = 2 if 'Coefficient' in df.columns else 2
        data = Reference(ws, min_col=importance_col, min_row=2, max_row=min(ws.max_row, 16), max_col=importance_col)
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 16))
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"F{ws.max_row + 3}")
    except:
        pass

def add_predictions_chart(ws, df):
    """
    Add predictions vs actual scatter chart
    """
    try:
        chart = ScatterChart()
        chart.title = "Predictions vs Actual"
        chart.style = 13
        chart.y_axis.title = 'Predicted'
        chart.x_axis.title = 'Actual'
        
        # Add scatter data
        x_data = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 100), max_col=1)
        y_data = Reference(ws, min_col=2, min_row=2, max_row=min(ws.max_row, 100), max_col=2)
        
        series = Series(y_data, x_data, title="Predictions")
        chart.series.append(series)
        
        ws.add_chart(chart, f"F{ws.max_row + 3}")
    except:
        pass

def add_missing_values_chart(ws, df):
    """
    Add missing values bar chart
    """
    try:
        chart = LineChart()
        chart.title = "Missing Values by Feature"
        chart.style = 13
        chart.y_axis.title = 'Missing Count'
        chart.x_axis.title = 'Features'
        
        # Add missing values data
        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"F{ws.max_row + 3}")
    except:
        pass

def create_summary_sheet(wb, results, timestamp):
    """
    Create a summary sheet with key insights
    """
    try:
        ws = wb.create_sheet(f"Summary_{timestamp}")
        
        # Add title
        ws['A1'] = "Excel ML Chat Assistant - Analysis Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Add timestamp
        ws['A2'] = f"Analysis completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        row = 4
        
        # Add key insights
        ws[f'A{row}'] = "Key Insights:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for sheet_name, df in results.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                ws[f'A{row}'] = f"• {sheet_name}: {len(df)} rows of analysis"
                row += 1
        
        # Format summary sheet
        for cell in ws['A1:A10']:
            if cell[0].value:
                cell[0].font = Font(size=11)
        
        print("✅ Created summary sheet")
        
    except Exception as e:
        print(f"⚠️ Could not create summary sheet: {str(e)}")

def export_visualizations_to_excel(file_path, viz_path="visualizations"):
    """
    Export visualization images to Excel
    """
    try:
        if not os.path.exists(viz_path):
            return "No visualizations found"
        
        wb = load_workbook(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ws = wb.create_sheet(f"Visualizations_{timestamp}")
        
        row = 1
        for filename in os.listdir(viz_path):
            if filename.endswith('.png'):
                img_path = os.path.join(viz_path, filename)
                img = Image(img_path)
                
                # Resize image
                img.width = 400
                img.height = 300
                
                # Add image to sheet
                ws.add_image(img, f'A{row}')
                ws[f'A{row}'] = filename.replace('.png', '').replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True)
                
                row += 20  # Space for next image
        
        wb.save(file_path)
        return f"Added visualizations to Excel"
        
    except Exception as e:
        return f"Error adding visualizations: {str(e)}"
